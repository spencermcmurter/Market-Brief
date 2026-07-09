"""
Builds the two things you receive: the .xlsx attachment and the HTML email body.
Headlines and the Link column are now real clickable hyperlinks to the article.
"""
import datetime as dt
import html

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

NAVY = PatternFill("solid", start_color="1F3B4D")
HIGH = PatternFill("solid", start_color="FBE3E1")
MED = PatternFill("solid", start_color="FFF3D6")
LOW = PatternFill("solid", start_color="EEF1F4")
LEG = PatternFill("solid", start_color="F4F6F8")
HFONT = Font(bold=True, color="FFFFFF", name="Arial", size=11)
BODY = Font(name="Arial", size=10)
BOLD = Font(name="Arial", size=10, bold=True)
LINK = Font(name="Arial", size=10, color="0563C1", underline="single")
LINKB = Font(name="Arial", size=10, color="0563C1", underline="single", bold=True)
TITLE = Font(name="Arial", size=13, bold=True, color="1F3B4D")
LEGF = Font(name="Arial", size=9, italic=True, color="555555")
THIN = Side(style="thin", color="D5DAE0")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
FILL = {"High": HIGH, "Medium": MED, "Low": LOW}


def _when(item):
    p = item.get("published")
    if isinstance(p, dt.datetime):
        return p.strftime("%b %d")
    return "recent"


def _star(item):
    if item.get("read_first"):
        return "\u2605"      # star
    if item.get("company_star"):
        return "*"
    return ""


def build_xlsx(sections, econ, earnings, path, date_label):
    wb = Workbook()
    ws = wb.active
    ws.title = f"Brief {date_label}"[:31]
    heads = ["Rank", "\u2605 / *", "Subject", "Exchange", "Category", "Headline",
             "Why it matters", "When", "Source", "Link"]
    widths = [8, 7, 20, 12, 20, 46, 48, 12, 18, 12]

    ws.merge_cells("A1:J1")
    ws["A1"] = f"Atlas \u2014 Daily Market Brief \u00b7 {date_label}"
    ws["A1"].font = TITLE
    ws.merge_cells("A2:J2")
    ws["A2"] = ("Legend:  \u2605 = read first   \u00b7   * = major event for that company   "
                "\u00b7   rows shaded by rank (red High / amber Medium / grey Low)   "
                "\u00b7   click a Headline or the Link to open the article")
    ws["A2"].font = LEGF
    ws["A2"].fill = LEG
    ws.row_dimensions[1].height = 20
    ws.row_dimensions[2].height = 16

    for c, (h, w) in enumerate(zip(heads, widths), 1):
        cell = ws.cell(3, c, h)
        cell.fill = NAVY
        cell.font = HFONT
        cell.border = BORDER
        cell.alignment = Alignment(vertical="center", wrap_text=True)
        ws.column_dimensions[cell.column_letter].width = w
    ws.row_dimensions[3].height = 28
    ws.freeze_panes = "A4"

    ordered = (sections["read_first"]
               + [x for x in sections["high"] if not x.get("read_first")]
               + sections["medium"] + sections["low"])
    r = 4
    for it in ordered:
        link = it.get("link", "")
        vals = [it["tier"], _star(it), it.get("subject", ""), it.get("exchange", "") or "\u2014",
                it.get("category", ""), it.get("title", ""), it.get("why", ""),
                _when(it), it.get("source", "") or "\u2014",
                "Open \u2197" if link else "\u2014"]
        f = FILL[it["tier"]]
        for c, v in enumerate(vals, 1):
            cell = ws.cell(r, c, v)
            cell.fill = f
            cell.border = BORDER
            cell.alignment = Alignment(vertical="top", wrap_text=True,
                                       horizontal="center" if c in (1, 2, 10) else "left")
            if c == 6 and link:                      # clickable headline
                cell.hyperlink = link
                cell.font = LINKB
            elif c == 10 and link:                   # clickable "Open" link
                cell.hyperlink = link
                cell.font = LINK
            elif c in (1, 2):
                cell.font = BOLD
            else:
                cell.font = BODY
        ws.row_dimensions[r].height = 44
        r += 1

    ws.merge_cells(start_row=r + 1, start_column=1, end_row=r + 1, end_column=10)
    note = ws.cell(r + 1, 1, "Filtered out: opinion / stock-tip content "
                            "(Motley Fool, Zacks, Seeking Alpha opinion, TipRanks, etc.)")
    note.font = LEGF
    note.fill = LEG

    cal = wb.create_sheet("Calendar")
    cal["A1"] = "Economic releases (US / CA, this week)"
    cal["A1"].font = TITLE
    row = 3
    for e in econ:
        cal.cell(row, 1, e.get("date", ""))
        cal.cell(row, 2, e.get("country", ""))
        cal.cell(row, 3, e.get("event", ""))
        cal.cell(row, 4, e.get("impact", ""))
        row += 1
    cal.column_dimensions["A"].width = 20
    cal.column_dimensions["C"].width = 50

    wb.save(path)
    return path


def build_email_html(sections, econ, date_label):
    def line(it):
        star = _star(it)
        prefix = f"{star} " if star else ""
        why = f" \u2014 {html.escape(it['why'])}" if it.get("why") else ""
        src = html.escape(it.get("source", "") or "")
        link = it.get("link", "")
        title = html.escape(it.get("title", ""))
        title_html = f'<a href="{link}">{title}</a>' if link else title
        return f"<li>{prefix}<b>{title_html}</b>{why} <span style='color:#888'>({src})</span></li>"

    parts = [f"<h2 style='font-family:Arial'>Atlas \u2014 Daily Market Brief \u00b7 {date_label}</h2>"]

    if sections["read_first"]:
        parts.append("<h3 style='font-family:Arial'>\u23f1 Read first (\u22485 min)</h3><ol style='font-family:Arial;font-size:14px'>")
        parts += [line(x) for x in sections["read_first"]]
        parts.append("</ol>")

    rest_high = [x for x in sections["high"] if not x.get("read_first")]
    if rest_high:
        parts.append("<h3 style='font-family:Arial'>\U0001f534 High impact</h3><ul style='font-family:Arial;font-size:14px'>")
        parts += [line(x) for x in rest_high]
        parts.append("</ul>")

    if sections["medium"]:
        parts.append("<h3 style='font-family:Arial'>\U0001f7e1 Medium</h3><ul style='font-family:Arial;font-size:14px'>")
        parts += [line(x) for x in sections["medium"]]
        parts.append("</ul>")

    if econ:
        parts.append("<h3 style='font-family:Arial'>\U0001f5d3 Calendar (US / CA)</h3><ul style='font-family:Arial;font-size:13px'>")
        for e in econ[:12]:
            parts.append(f"<li>{html.escape(e.get('date',''))} \u00b7 {html.escape(e.get('country',''))} \u00b7 "
                         f"{html.escape(e.get('event',''))} <i>({html.escape(e.get('impact',''))})</i></li>")
        parts.append("</ul>")

    n = len(sections["high"]) + len(sections["medium"]) + len(sections["low"])
    parts.append(f"<p style='font-family:Arial;color:#888;font-size:12px'>Full ranked list attached "
                 f"({n} items; tip/opinion content filtered).</p>")
    return "".join(parts)
